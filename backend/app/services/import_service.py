"""Excel / .xlsm bulk import (blueprint §12-F: 'Import Existing Excel').

Expected workbook (case-insensitive sheet + column names; whitespace tolerant):

  Sheet 'students' (optional):
    full_name | gender         | guardian_name | guardian_phone
    Ali       | MALE | M | F   | Father        | 0501234567

  Sheet 'progress' (optional):
    student_full_name | surah_order | status            | completion_percent
    Ali               | 1           | MASTERED          | 100

Both sheets are optional but at least one must be present. Students are
matched (case-insensitive) by full_name within the caller's school; if no
match, a new student is created. Progress rows reference an already-imported
or pre-existing student by name.

Per-row errors are collected and returned; one bad row never aborts the
whole import.
"""
from __future__ import annotations

from io import BytesIO
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.audit_log import AuditAction, AuditEntityType
from app.models.memorization_progress import MemorizationProgress, MemorizationStatus
from app.models.progress_history import ProgressHistory
from app.models.student import Student, StudentGender
from app.models.surah import QuranSurah
from app.repositories import progress_repo, surah_repo
from app.schemas.import_ import ImportError as ImportErrorSchema
from app.schemas.import_ import ImportResult
from app.services import audit_service

# Map common column header variants to our canonical fields. Lowercase keys.
_STUDENT_ALIASES = {
    "full_name": "full_name",
    "name": "full_name",
    "student": "full_name",
    "student_name": "full_name",
    "gender": "gender",
    "sex": "gender",
    "guardian_name": "guardian_name",
    "guardian": "guardian_name",
    "parent_name": "guardian_name",
    "guardian_phone": "guardian_phone",
    "phone": "guardian_phone",
    "parent_phone": "guardian_phone",
}

_PROGRESS_ALIASES = {
    "student_full_name": "student_full_name",
    "student_name": "student_full_name",
    "student": "student_full_name",
    "name": "student_full_name",
    "surah_order": "surah_order",
    "surah_no": "surah_order",
    "surah": "surah_order",
    "status": "status",
    "memorization_status": "status",
    "completion_percent": "completion_percent",
    "percent": "completion_percent",
    "completion": "completion_percent",
}

_GENDER_ALIASES = {
    "male": StudentGender.MALE,
    "m": StudentGender.MALE,
    "boy": StudentGender.MALE,
    "female": StudentGender.FEMALE,
    "f": StudentGender.FEMALE,
    "girl": StudentGender.FEMALE,
}


def _norm(s: Any) -> str:
    return str(s).strip().lower() if s is not None else ""


def _build_header_map(headers: list[Any], aliases: dict[str, str]) -> dict[str, int]:
    """Return canonical_field -> column_index for the headers we recognize."""
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(headers):
        key = _norm(raw).replace(" ", "_")
        canonical = aliases.get(key)
        if canonical and canonical not in mapping:
            mapping[canonical] = idx
    return mapping


async def import_workbook(
    db: AsyncSession,
    *,
    actor_id: UUID,
    school_id: UUID,
    file_bytes: bytes,
) -> ImportResult:
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        return ImportResult(
            errors=[ImportErrorSchema(sheet="-", row=0, message=f"Cannot open workbook: {e}")]
        )

    # Lowercase sheet name lookup so 'Students' / 'STUDENTS' / 'students' all match.
    sheets_by_name = {name.lower(): name for name in wb.sheetnames}
    if "students" not in sheets_by_name and "progress" not in sheets_by_name:
        return ImportResult(
            errors=[
                ImportErrorSchema(
                    sheet="-",
                    row=0,
                    message="Workbook must contain at least a 'students' or 'progress' sheet.",
                )
            ]
        )

    result = ImportResult()
    name_to_student: dict[str, Student] = {}

    # Pre-load existing students in the school for name-based matching.
    existing_stmt = select(Student).where(Student.school_id == school_id)
    for s in (await db.execute(existing_stmt)).scalars().all():
        name_to_student[s.full_name.strip().lower()] = s

    if "students" in sheets_by_name:
        await _import_students(
            db,
            actor_id=actor_id,
            school_id=school_id,
            sheet=wb[sheets_by_name["students"]],
            result=result,
            name_to_student=name_to_student,
        )

    if "progress" in sheets_by_name:
        # Pre-load surahs by surah_order for reference.
        surahs_by_order: dict[int, QuranSurah] = {
            s.surah_order: s for s in await surah_repo.list_all(db)
        }
        await _import_progress(
            db,
            actor_id=actor_id,
            school_id=school_id,
            sheet=wb[sheets_by_name["progress"]],
            result=result,
            name_to_student=name_to_student,
            surahs_by_order=surahs_by_order,
        )

    return result


async def _import_students(
    db: AsyncSession,
    *,
    actor_id: UUID,
    school_id: UUID,
    sheet,
    result: ImportResult,
    name_to_student: dict[str, Student],
) -> None:
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return
    cols = _build_header_map(list(headers), _STUDENT_ALIASES)

    if "full_name" not in cols:
        result.errors.append(
            ImportErrorSchema(
                sheet="students",
                row=1,
                message="Missing required column 'full_name' (or 'name').",
            )
        )
        return

    for row_idx, row in enumerate(rows, start=2):
        full_name_val = row[cols["full_name"]] if cols["full_name"] < len(row) else None
        full_name = str(full_name_val).strip() if full_name_val else ""
        if not full_name:
            continue  # silently skip blank rows

        key = full_name.lower()
        if key in name_to_student:
            result.students_matched += 1
            continue

        gender_idx = cols.get("gender")
        gender_val = (
            row[gender_idx] if gender_idx is not None and gender_idx < len(row) else None
        )
        gender = _GENDER_ALIASES.get(_norm(gender_val))
        if gender is None:
            result.errors.append(
                ImportErrorSchema(
                    sheet="students",
                    row=row_idx,
                    message=f"Invalid or missing gender for '{full_name}' "
                    "(expected MALE/FEMALE/M/F).",
                )
            )
            continue

        guardian_name = _opt_str(row, cols.get("guardian_name"))
        guardian_phone = _opt_str(row, cols.get("guardian_phone"))

        student = Student(
            school_id=school_id,
            full_name=full_name,
            gender=gender,
            guardian_name=guardian_name,
            guardian_phone=guardian_phone,
        )
        db.add(student)
        await db.flush()
        await db.refresh(student)
        name_to_student[key] = student
        result.students_created += 1

        await audit_service.record(
            db,
            actor_id=actor_id,
            school_id=school_id,
            action=AuditAction.CREATE,
            entity_type=AuditEntityType.STUDENT,
            entity_id=student.id,
            new_value=audit_service.snapshot(
                student,
                ("full_name", "gender", "guardian_name", "guardian_phone", "school_id"),
            ),
        )


async def _import_progress(
    db: AsyncSession,
    *,
    actor_id: UUID,
    school_id: UUID,
    sheet,
    result: ImportResult,
    name_to_student: dict[str, Student],
    surahs_by_order: dict[int, QuranSurah],
) -> None:
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration:
        return
    cols = _build_header_map(list(headers), _PROGRESS_ALIASES)

    missing = [
        c for c in ("student_full_name", "surah_order", "status") if c not in cols
    ]
    if missing:
        result.errors.append(
            ImportErrorSchema(
                sheet="progress",
                row=1,
                message=f"Missing required column(s): {', '.join(missing)}",
            )
        )
        return

    progress_audit_fields = (
        "student_id",
        "surah_id",
        "status",
        "completion_percent",
        "teacher_id",
    )

    for row_idx, row in enumerate(rows, start=2):
        name_val = row[cols["student_full_name"]] if cols["student_full_name"] < len(row) else None
        full_name = str(name_val).strip() if name_val else ""
        if not full_name:
            continue

        student = name_to_student.get(full_name.lower())
        if student is None:
            result.errors.append(
                ImportErrorSchema(
                    sheet="progress",
                    row=row_idx,
                    message=f"No student named '{full_name}' (add them to the students "
                    "sheet or create them first).",
                )
            )
            continue

        # surah_order
        surah_val = row[cols["surah_order"]] if cols["surah_order"] < len(row) else None
        try:
            surah_order = int(surah_val) if surah_val is not None else None
        except (TypeError, ValueError):
            surah_order = None
        surah = surahs_by_order.get(surah_order) if surah_order else None
        if surah is None:
            result.errors.append(
                ImportErrorSchema(
                    sheet="progress",
                    row=row_idx,
                    message=f"Invalid surah_order {surah_val!r} (must be 1..114).",
                )
            )
            continue

        # status
        status_val = row[cols["status"]] if cols["status"] < len(row) else None
        status_str = _norm(status_val).upper().replace(" ", "_")
        try:
            status = MemorizationStatus(status_str) if status_str else None
        except ValueError:
            status = None
        if status is None:
            result.errors.append(
                ImportErrorSchema(
                    sheet="progress",
                    row=row_idx,
                    message=f"Invalid status {status_val!r}. Valid: "
                    + ", ".join(s.value for s in MemorizationStatus),
                )
            )
            continue

        # completion_percent (optional)
        pct_idx = cols.get("completion_percent")
        pct_val = row[pct_idx] if pct_idx is not None and pct_idx < len(row) else None
        try:
            completion = int(pct_val) if pct_val not in (None, "") else 0
        except (TypeError, ValueError):
            completion = 0
        completion = max(0, min(100, completion))

        existing = await progress_repo.get_by_student_and_surah(
            db, student.id, surah.id
        )
        action = AuditAction.UPDATE if existing else AuditAction.CREATE
        if existing is None:
            existing = MemorizationProgress(
                student_id=student.id,
                surah_id=surah.id,
                teacher_id=actor_id,
                status=status,
                completion_percent=completion,
            )
            db.add(existing)
        else:
            existing.status = status
            existing.completion_percent = completion
            existing.teacher_id = actor_id
        await db.flush()
        await db.refresh(existing)

        # Append history snapshot + audit log to keep import-driven changes
        # equivalent to API-driven ones.
        db.add(
            ProgressHistory(
                progress_id=existing.id,
                student_id=existing.student_id,
                surah_id=existing.surah_id,
                teacher_id=existing.teacher_id,
                status=existing.status,
                score=existing.score,
                completion_percent=existing.completion_percent,
                notes=existing.notes,
                action=action,
            )
        )
        await audit_service.record(
            db,
            actor_id=actor_id,
            school_id=school_id,
            action=action,
            entity_type=AuditEntityType.PROGRESS,
            entity_id=existing.id,
            new_value=audit_service.snapshot(existing, progress_audit_fields),
        )

        result.progress_recorded += 1


def _opt_str(row: tuple, idx: int | None) -> str | None:
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    s = str(val).strip()
    return s or None
