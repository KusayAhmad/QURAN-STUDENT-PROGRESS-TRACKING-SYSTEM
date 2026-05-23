"""Excel bulk import endpoint."""
from io import BytesIO

import pytest
from openpyxl import Workbook

pytestmark = pytest.mark.asyncio

# Reuse admin_client fixture
pytest_plugins = ["tests.test_audit_log"]


def _xlsx(students_rows: list[list] | None, progress_rows: list[list] | None) -> bytes:
    wb = Workbook()
    # Replace default sheet
    wb.remove(wb.active)
    if students_rows is not None:
        ws = wb.create_sheet("students")
        for row in students_rows:
            ws.append(row)
    if progress_rows is not None:
        ws = wb.create_sheet("progress")
        for row in progress_rows:
            ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _post_xlsx(admin_client, file_bytes: bytes):
    return await admin_client.post(
        "/api/v1/admin/import",
        files={
            "file": (
                "import.xlsx",
                file_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )


async def test_import_students_only(admin_client):
    data = _xlsx(
        students_rows=[
            ["full_name", "gender", "guardian_name", "guardian_phone"],
            ["Imported Ali", "MALE", "Father Ali", "0500000001"],
            ["Imported Fatima", "F", None, None],
        ],
        progress_rows=None,
    )
    res = await _post_xlsx(admin_client, data)
    assert res.status_code == 200, res.text
    body = res.json()
    assert body["students_created"] == 2
    assert body["students_matched"] == 0
    assert body["progress_recorded"] == 0
    assert body["errors"] == []

    # Verify visible via /students
    listing = await admin_client.get("/api/v1/students")
    names = [s["full_name"] for s in listing.json()["items"]]
    assert "Imported Ali" in names
    assert "Imported Fatima" in names


async def test_import_progress_matches_existing_student(admin_client, seeded_surahs):
    # Pre-create a student
    sid = (
        await admin_client.post(
            "/api/v1/students", json={"full_name": "Existing Bilal", "gender": "MALE"}
        )
    ).json()["id"]

    data = _xlsx(
        students_rows=None,
        progress_rows=[
            ["student_full_name", "surah_order", "status", "completion_percent"],
            ["Existing Bilal", 1, "MASTERED", 100],
            ["Existing Bilal", 2, "WEAK", 30],
        ],
    )
    res = await _post_xlsx(admin_client, data)
    body = res.json()
    assert body["progress_recorded"] == 2
    assert body["errors"] == []

    progress = await admin_client.get(f"/api/v1/students/{sid}/progress")
    rows = progress.json()
    by_surah = {r["surah_id"]: r for r in rows}
    assert by_surah[1]["status"] == "MASTERED"
    assert by_surah[2]["status"] == "WEAK"


async def test_import_creates_progress_history_entries(admin_client, seeded_surahs):
    """Imported progress writes to progress_history just like API writes."""
    data = _xlsx(
        students_rows=[["full_name", "gender"], ["History Test", "MALE"]],
        progress_rows=[
            ["student_full_name", "surah_order", "status"],
            ["History Test", 1, "STRONG"],
        ],
    )
    await _post_xlsx(admin_client, data)
    sid = next(
        s["id"]
        for s in (await admin_client.get("/api/v1/students")).json()["items"]
        if s["full_name"] == "History Test"
    )
    timeline = await admin_client.get(f"/api/v1/students/{sid}/surahs/1/timeline")
    body = timeline.json()
    assert len(body) == 1
    assert body[0]["status"] == "STRONG"
    assert body[0]["action"] == "CREATE"


async def test_import_combined_students_and_progress(admin_client, seeded_surahs):
    """Single workbook with both sheets — student created in pass 1 is then
    referenced by progress in pass 2."""
    data = _xlsx(
        students_rows=[
            ["name", "gender"],
            ["Combo Student", "MALE"],
        ],
        progress_rows=[
            ["student", "surah_order", "status", "completion"],
            ["Combo Student", 1, "MASTERED", 100],
        ],
    )
    res = await _post_xlsx(admin_client, data)
    body = res.json()
    assert body["students_created"] == 1
    assert body["progress_recorded"] == 1
    assert body["errors"] == []


async def test_import_duplicate_student_is_matched_not_recreated(admin_client):
    await admin_client.post(
        "/api/v1/students", json={"full_name": "Dup Student", "gender": "MALE"}
    )
    data = _xlsx(
        students_rows=[
            ["full_name", "gender"],
            ["Dup Student", "MALE"],
            ["DUP STUDENT", "MALE"],  # case-insensitive match -> still matched
        ],
        progress_rows=None,
    )
    res = await _post_xlsx(admin_client, data)
    body = res.json()
    assert body["students_created"] == 0
    assert body["students_matched"] == 2


async def test_import_invalid_status_collected_as_error(admin_client, seeded_surahs):
    await admin_client.post(
        "/api/v1/students", json={"full_name": "Err Student", "gender": "MALE"}
    )
    data = _xlsx(
        students_rows=None,
        progress_rows=[
            ["student", "surah_order", "status"],
            ["Err Student", 1, "MASTERED"],   # ok
            ["Err Student", 2, "BOGUS"],      # bad status
            ["No Such Person", 1, "MASTERED"],  # bad name
            ["Err Student", 999, "STRONG"],   # bad surah_order
        ],
    )
    res = await _post_xlsx(admin_client, data)
    body = res.json()
    assert body["progress_recorded"] == 1
    assert len(body["errors"]) == 3
    messages = " ".join(e["message"] for e in body["errors"])
    assert "BOGUS" in messages or "Invalid status" in messages
    assert "No Such Person" in messages
    assert "999" in messages


async def test_import_missing_required_column(admin_client):
    data = _xlsx(
        students_rows=[["gender"], ["MALE"]],  # no full_name column
        progress_rows=None,
    )
    res = await _post_xlsx(admin_client, data)
    body = res.json()
    assert body["students_created"] == 0
    assert any("full_name" in e["message"] for e in body["errors"])


async def test_import_admin_only(auth_client):
    data = _xlsx(
        students_rows=[["full_name", "gender"], ["X", "MALE"]],
        progress_rows=None,
    )
    res = await auth_client.post(
        "/api/v1/admin/import",
        files={"file": ("x.xlsx", data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert res.status_code == 403


async def test_import_garbage_file(admin_client):
    res = await _post_xlsx(admin_client, b"this is not an xlsx")
    body = res.json()
    assert body["students_created"] == 0
    assert body["progress_recorded"] == 0
    assert body["errors"]
    assert "Cannot open" in body["errors"][0]["message"]


async def test_import_workbook_with_no_known_sheets(admin_client):
    wb = Workbook()
    wb.active.title = "RandomData"
    wb.active.append(["foo", "bar"])
    buf = BytesIO()
    wb.save(buf)
    res = await _post_xlsx(admin_client, buf.getvalue())
    body = res.json()
    assert body["errors"]
    assert "students" in body["errors"][0]["message"]
